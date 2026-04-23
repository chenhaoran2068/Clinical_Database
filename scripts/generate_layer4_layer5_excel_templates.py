from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
README_FILL = PatternFill("solid", fgColor="F4F7FB")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def add_readme_sheet(workbook: Workbook, title: str, lines: list[str]) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet["A1"] = "使用说明"
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].fill = README_FILL
    sheet["A2"] = "本模板用于初始化 Layer 4/Layer 5 的人工审查与机器可读注册表。字段名应保持稳定，说明文本可按数据库实际情况补充。"
    sheet["A2"].alignment = WRAP_ALIGNMENT
    row = 4
    for line in lines:
        sheet.cell(row=row, column=1, value=line)
        sheet.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
        row += 1
    sheet.column_dimensions["A"].width = 120


def add_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[tuple[str, str]],
    rows: list[list[object]] | None = None,
) -> None:
    sheet = workbook.create_sheet(title=title)
    for index, (header, comment_text) in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT
        if comment_text:
            cell.comment = Comment(comment_text, "Codex")
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 4, 16), 40)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    if rows:
        for row_index, row_values in enumerate(rows, start=2):
            for column_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = WRAP_ALIGNMENT


def build_layer4_global_template(output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_readme_sheet(
        workbook,
        "README",
        [
            "一行代表一个数据库版本。",
            "database_id 应保持跨文档一致，例如 MIMIC-IV-3.1、AmsterdamUMCdb-1.0.2。",
            "current_status 建议使用 planned / received / profiled / layer2_active / layer3_active / archived。",
            "local_layer*_path 建议填写绝对路径或从仓库根目录可解析的相对路径。",
        ],
    )
    add_table_sheet(
        workbook,
        "database_catalog",
        [
            ("database_type", "数据库大类，例如 clinical_icu_ehr、hospital_ehr、survey、claims。"),
            ("database_id", "数据库稳定标识符。"),
            ("database_name", "数据库公开名称。"),
            ("version", "版本号或快照标记。"),
            ("acquisition_date", "获取或接收日期。"),
            ("source_org", "数据提供方或维护机构。"),
            ("source_website", "官方主页或下载说明链接。"),
            ("license_scope", "许可范围与使用限制摘要。"),
            ("local_layer1_path", "该数据库在 Layer 1 的主目录。"),
            ("local_layer2_path", "该数据库在 Layer 2 的主目录。"),
            ("local_layer3_scope", "该数据库当前涉及的 Layer 3 语义范围。"),
            ("registry_owner", "当前负责维护注册表的人。"),
            ("current_status", "当前整体状态。"),
            ("last_update_at", "最后更新时间。"),
            ("remarks", "补充说明。"),
        ],
    )

    ensure_parent(output_path)
    workbook.save(output_path)


def build_layer4_per_database_template(output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_readme_sheet(
        workbook,
        "README",
        [
            "本模板对应单个数据库的 Source Registry。",
            "raw_tables / raw_columns 用于登记源表和源字段的结构、键、时间语义与审查状态。",
            "layer2_assets / derivation_rules / time_anchor_rules 用于记录 Layer 2 的拆分、清洗、派生与时间锚策略。",
            "建议先完成 raw_tables，再逐步补 raw_columns。大型分片表可在 physical_path 中记录通配符和文件数摘要。",
        ],
    )
    add_table_sheet(
        workbook,
        "database_overview",
        [
            ("database_id", "数据库稳定标识符。"),
            ("database_name", "数据库名称。"),
            ("version", "版本号。"),
            ("database_type", "数据库大类。"),
            ("source_org", "提供机构。"),
            ("download_or_receive_date", "接收日期。"),
            ("layer1_root", "Layer 1 根目录。"),
            ("primary_timezone", "主要时间语义所属时区。"),
            ("primary_patient_key", "主要 patient-level 键。"),
            ("primary_encounter_key", "主要 admission / encounter 键。"),
            ("primary_stay_key", "主要 ICU stay 键。"),
            ("data_license", "许可与使用限制。"),
            ("overall_notes", "数据库级总体说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "raw_tables",
        [
            ("raw_table_id", "逻辑源表标识，例如 hosp.labevents。"),
            ("raw_table_name", "源表名称。"),
            ("source_domain", "来源域，例如 hosp / icu / ed / note / survey。"),
            ("physical_path", "Layer 1 中对应文件或通配路径。"),
            ("original_format", "原始格式，例如 csv / parquet / xpt。"),
            ("grain_level", "记录粒度摘要。"),
            ("row_count_estimate", "行数估计。"),
            ("primary_keys", "主键或候选主键。"),
            ("patient_key_cols", "患者键字段。"),
            ("encounter_key_cols", "住院或 encounter 键字段。"),
            ("stay_key_cols", "stay 键字段。"),
            ("time_columns", "时间列。"),
            ("unit_summary", "单位结构摘要。"),
            ("derived_flag", "是否为衍生表。"),
            ("upstream_source_hint", "若为衍生表，记录上游来源线索。"),
            ("review_status", "审查状态。"),
            ("reviewed_by", "审查人。"),
            ("reviewed_at", "审查时间。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "raw_columns",
        [
            ("raw_table_id", "所属逻辑源表。"),
            ("column_name", "字段名。"),
            ("logical_name_cn", "中文逻辑名。"),
            ("source_declared_type", "官方说明中的字段类型。"),
            ("data_type_raw", "Layer 1 观测到的物理类型。"),
            ("data_type_normalized", "归一化后的数据类型。"),
            ("schema_nullable_flag", "schema 是否允许为空。"),
            ("observed_null_flag", "抽样是否观测到空值。"),
            ("unit", "单位或单位来源提示。"),
            ("code_system", "编码体系提示。"),
            ("example_values", "示例值。"),
            ("missing_pattern", "缺失模式摘要。"),
            ("value_range_or_levels", "值域或枚举水平。"),
            ("distribution_summary", "抽样分布摘要。"),
            ("is_key", "是否承担键角色。"),
            ("key_role", "键角色，如 patient / encounter / stay / event。"),
            ("is_time", "是否为时间列。"),
            ("time_semantics", "时间语义标签。"),
            ("derived_flag", "字段是否属于衍生表。"),
            ("derived_from", "字段来源线索。"),
            ("review_status", "审查状态。"),
            ("reviewed_by", "审查人。"),
            ("reviewed_at", "审查时间。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "layer2_assets",
        [
            ("layer2_asset_id", "Layer 2 资产标识。"),
            ("asset_name", "资产名称。"),
            ("asset_type", "资产类型，例如 reviewed_unsplit / split_cleaned / time_anchor。"),
            ("source_raw_table_ids", "来源 raw_table_id 列表。"),
            ("output_path", "资产输出路径。"),
            ("grain_level", "资产粒度。"),
            ("patient_key_cols", "患者键字段。"),
            ("encounter_key_cols", "住院或 encounter 键字段。"),
            ("stay_key_cols", "stay 键字段。"),
            ("time_anchor_id", "关联的 time anchor 规则。"),
            ("split_required", "是否需要分片。"),
            ("cleaning_scope", "清洗范围。"),
            ("process_batch_id", "处理批次编号。"),
            ("status", "状态。"),
            ("owner", "负责人。"),
            ("updated_at", "更新时间。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "derivation_rules",
        [
            ("rule_id", "规则编号。"),
            ("rule_name", "规则名称。"),
            ("input_tables", "输入表或 Layer 2 资产。"),
            ("output_assets", "输出资产。"),
            ("trigger_condition", "触发条件。"),
            ("split_logic", "拆分逻辑。"),
            ("cleaning_logic", "清洗逻辑。"),
            ("null_handling", "缺失值处理。"),
            ("unit_handling", "单位处理。"),
            ("duplicate_handling", "重复处理。"),
            ("approval_status", "审批状态。"),
            ("approved_by", "审批人。"),
            ("approved_at", "审批时间。"),
            ("version", "规则版本。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "time_anchor_rules",
        [
            ("anchor_id", "时间锚标识。"),
            ("anchor_name", "时间锚名称。"),
            ("anchor_level", "锚点层级，例如 patient / encounter / stay。"),
            ("source_table", "锚点来源表。"),
            ("source_time_column", "锚点时间字段。"),
            ("join_keys", "关联键。"),
            ("anchor_definition", "锚点定义。"),
            ("timezone_rule", "时区处理规则。"),
            ("relative_time_unit", "相对时间单位，例如 minutes / hours / days。"),
            ("negative_time_allowed", "是否允许负时间。"),
            ("output_path", "锚点资产输出路径。"),
            ("approval_status", "审批状态。"),
            ("approved_by", "审批人。"),
            ("approved_at", "审批时间。"),
            ("remarks", "补充说明。"),
        ],
    )

    ensure_parent(output_path)
    workbook.save(output_path)


def build_layer5_global_template(output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_readme_sheet(
        workbook,
        "README",
        [
            "一行代表一个标准变量。",
            "std_variable_id 必须全局稳定，且跨数据库保持同名即同义。",
            "semantic_folder_dictionary 是 Layer 5 全局语义分区字典，可作为 Layer 3/5 新变量的默认落点参考。",
        ],
    )
    add_table_sheet(
        workbook,
        "std_variable_catalog",
        [
            ("std_variable_id", "标准变量稳定标识符。"),
            ("std_variable_name_cn", "中文名称。"),
            ("std_variable_name_en", "英文名称。"),
            ("semantic_folder", "语义目录。"),
            ("definition", "标准定义。"),
            ("value_type", "值类型。"),
            ("standard_unit", "标准单位。"),
            ("record_grain", "记录粒度。"),
            ("default_anchor_type", "默认时间锚类型。"),
            ("active_databases", "当前已接入数据库列表。"),
            ("current_status", "当前状态，例如 active / draft / retired。"),
            ("materialization_status", "物化状态，例如 retained / pending_build。"),
            ("latest_process_batch_id", "最近处理批次。"),
            ("current_row_count", "当前行数。"),
            ("layer3_asset_path", "Layer 3 资产路径。"),
            ("knowledge_package_path", "Layer 5 知识包路径。"),
            ("log_archive_path", "日志归档路径。"),
            ("owner", "负责人。"),
            ("latest_version", "当前版本。"),
            ("latest_review_date", "最近审查日期。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "semantic_folder_dictionary",
        [
            ("semantic_folder", "语义目录代号。"),
            ("display_name_cn", "中文显示名。"),
            ("scope_note", "适用范围说明。"),
        ],
        rows=[
            ["id_mapping", "ID 映射", "患者、住院、stay 等主键映射与桥接资产。"],
            ["encounter_information", "就诊信息", "入院、出院、ICU stay 来源与去向等结构信息。"],
            ["demographics", "人口学", "年龄、性别、种族等。"],
            ["anthropometrics", "体格测量", "身高、体重、BMI 等。"],
            ["vital_signs", "生命体征", "心率、血压、体温、呼吸频率、SpO2 等。"],
            ["laboratory", "实验室", "检验结果与化验相关变量。"],
            ["diagnosis_history", "既往诊断", "历史疾病与病史相关变量。"],
            ["diagnosis_current", "当前诊断", "本次住院/ICU 当前诊断。"],
            ["diagnosis_computable", "可计算诊断", "由规则推导的表型或诊断事件。"],
            ["orders", "医嘱", "下达的检查、治疗、处方等医嘱。"],
            ["medication", "用药", "药物使用、给药证据与成分映射。"],
            ["nursing_execution_or_documented_care", "护理执行/护理记录", "护理执行事件或文书化护理。"],
            ["treatment_intervention", "治疗干预", "操作、治疗和干预事件。"],
            ["treatment_state", "治疗状态", "治疗是否进行中的状态型变量。"],
            ["device_support", "设备支持", "呼吸机、ECMO、CRRT 等设备支持。"],
            ["supportive_therapy", "支持治疗", "氧疗、机械通气、血流动力支持等。"],
            ["scores", "评分", "SOFA、APS、GCS 等评分。"],
            ["outcomes", "结局", "死亡、LOS、出院去向等结局。"],
            ["intake_output_balance", "出入量", "入量、出量、平衡等液体相关资产。"],
        ],
    )

    ensure_parent(output_path)
    workbook.save(output_path)


def build_layer5_per_variable_template(output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_readme_sheet(
        workbook,
        "README",
        [
            "本模板对应单个标准变量的完整知识包。",
            "variable_profile 记录该变量当前版本的标准定义、路径与状态。",
            "standard_contract 记录跨数据库必须保持一致的核心规范。",
            "database_notes / issue_resolution_log / reprocessing_assessment 用于存放数据库特有差异、问题闭环和重处理评估。",
        ],
    )
    add_table_sheet(
        workbook,
        "variable_profile",
        [
            ("std_variable_id", "标准变量标识。"),
            ("std_variable_name_cn", "中文名称。"),
            ("std_variable_name_en", "英文名称。"),
            ("semantic_folder", "语义目录。"),
            ("definition", "标准定义。"),
            ("value_type", "值类型。"),
            ("standard_unit", "标准单位。"),
            ("record_grain", "记录粒度。"),
            ("default_anchor_type", "默认时间锚。"),
            ("layer3_asset_path", "Layer 3 资产路径。"),
            ("layer5_manifest_path", "Layer 5 manifest 路径。"),
            ("layer5_preview_path", "Layer 5 preview 路径。"),
            ("current_status", "当前状态。"),
            ("materialization_status", "物化状态。"),
            ("latest_process_batch_id", "最近处理批次。"),
            ("current_row_count", "当前行数。"),
            ("owner", "负责人。"),
            ("current_version", "当前版本。"),
            ("summary_note", "摘要说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "standard_contract",
        [
            ("rule_order", "规则顺序。"),
            ("rule_dimension", "规则维度，例如 unit / value_domain / precision / merge_rule。"),
            ("target_field", "作用字段。"),
            ("canonical_rule", "规范要求。"),
            ("allowed_values_or_range", "允许值域或范围。"),
            ("source_mapping_policy", "源字段映射策略。"),
            ("applies_to_scope", "适用范围。"),
            ("if_unmappable_then", "无法映射时的处理方式。"),
            ("change_requires_reprocessing_assessment", "规则变更是否需要重处理评估。"),
            ("approved_by", "审批人。"),
            ("approved_at", "审批时间。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "layer3_schema",
        [
            ("column_order", "列顺序。"),
            ("column_name", "Layer 3 字段名。"),
            ("column_name_cn", "中文字段名。"),
            ("data_type", "数据类型。"),
            ("nullable_flag", "是否可空。"),
            ("definition", "字段定义。"),
            ("unit", "单位。"),
            ("precision_rule", "精度规则。"),
            ("key_role", "键角色。"),
            ("allowed_values_or_range", "允许值域或范围。"),
            ("usage_caution", "使用注意事项。"),
        ],
    )
    add_table_sheet(
        workbook,
        "global_cautions",
        [
            ("caution_id", "警示编号。"),
            ("caution_level", "警示级别。"),
            ("topic", "主题。"),
            ("description", "详细说明。"),
            ("trigger_condition", "触发条件。"),
            ("preferred_handling", "推荐处理方式。"),
            ("added_by", "登记人。"),
            ("added_at", "登记时间。"),
            ("active_flag", "当前是否仍有效。"),
            ("superseded_by", "若已替代，记录新警示编号。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "database_notes",
        [
            ("database_id", "数据库标识。"),
            ("source_tables", "相关源表。"),
            ("source_codes_or_columns", "相关源字段或编码。"),
            ("issue_summary", "数据库特有问题摘要。"),
            ("handling_strategy", "处理策略。"),
            ("unresolved_risk", "剩余风险。"),
            ("needs_recheck_flag", "是否需要复查。"),
            ("related_process_batch_id", "相关批次编号。"),
            ("recorded_by", "记录人。"),
            ("recorded_at", "记录时间。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "issue_resolution_log",
        [
            ("issue_id", "问题编号。"),
            ("database_id", "所属数据库。"),
            ("discovered_in_batch", "发现批次。"),
            ("problem_type", "问题类型。"),
            ("problem_description", "问题描述。"),
            ("impact_scope", "影响范围。"),
            ("resolution_decision", "处理决策。"),
            ("resolution_status", "处理状态。"),
            ("resolved_by", "处理人。"),
            ("resolved_at", "处理时间。"),
            ("linked_log_id", "相关日志编号。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "reprocessing_assessment",
        [
            ("assessment_id", "评估编号。"),
            ("trigger_source", "触发来源。"),
            ("affected_databases", "受影响数据库。"),
            ("affected_layer3_assets", "受影响 Layer 3 资产。"),
            ("reprocessing_required", "是否需要重处理。"),
            ("priority", "优先级。"),
            ("reason", "原因。"),
            ("decision_owner", "决策人。"),
            ("decision_date", "决策日期。"),
            ("completion_status", "完成状态。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "log_index",
        [
            ("log_id", "日志编号。"),
            ("process_batch_id", "批次编号。"),
            ("database_id", "数据库标识。"),
            ("log_type", "日志类型。"),
            ("log_date", "日志日期。"),
            ("processed_by", "执行人。"),
            ("reviewed_by", "审查人。"),
            ("file_path", "日志文件路径。"),
            ("supersedes_log_id", "若替代旧日志，记录旧编号。"),
            ("signature_status", "签名或确认状态。"),
            ("remarks", "补充说明。"),
        ],
    )
    add_table_sheet(
        workbook,
        "change_history",
        [
            ("version_id", "版本编号。"),
            ("change_date", "变更日期。"),
            ("change_type", "变更类型。"),
            ("change_summary", "变更摘要。"),
            ("author", "作者。"),
            ("reviewer", "审查人。"),
            ("effective_from", "生效时间。"),
            ("remarks", "补充说明。"),
        ],
    )

    ensure_parent(output_path)
    workbook.save(output_path)


def main() -> None:
    repo_root = Path(__file__).resolve().parents[3]
    layer4_dir = repo_root / "Methods" / "Clinical_Database" / "local_work" / "Layer 4" / "Templates"
    layer5_dir = repo_root / "Methods" / "Clinical_Database" / "local_work" / "Layer 5" / "Templates"

    build_layer4_global_template(layer4_dir / "Layer4_Global_DatabaseCatalog_Template.xlsx")
    build_layer4_per_database_template(layer4_dir / "Layer4_PerDatabase_SourceRegistry_Template.xlsx")
    build_layer5_global_template(layer5_dir / "Layer5_StdVariable_MasterIndex_Template.xlsx")
    build_layer5_per_variable_template(layer5_dir / "Layer5_PerVariable_KnowledgePackage_Template.xlsx")

    print("Generated Layer 4 and Layer 5 Excel templates:")
    print(layer4_dir / "Layer4_Global_DatabaseCatalog_Template.xlsx")
    print(layer4_dir / "Layer4_PerDatabase_SourceRegistry_Template.xlsx")
    print(layer5_dir / "Layer5_StdVariable_MasterIndex_Template.xlsx")
    print(layer5_dir / "Layer5_PerVariable_KnowledgePackage_Template.xlsx")


if __name__ == "__main__":
    main()
