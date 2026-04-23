from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[4]

MASTER_WORKBOOKS = [
    REPO_ROOT / "Methods" / "Clinical_Database" / "local_work" / "Layer 5" / "Global" / "Layer5_StdVariable_MasterIndex.xlsx",
    REPO_ROOT / "Methods" / "Clinical_Database" / "local_work" / "Layer 5" / "Templates" / "Layer5_StdVariable_MasterIndex_Template.xlsx",
]

PER_VARIABLE_TEMPLATE = REPO_ROOT / "Methods" / "Clinical_Database" / "local_work" / "Layer 5" / "Templates" / "Layer5_PerVariable_KnowledgePackage_Template.xlsx"

MASTER_HEADERS = [
    "std_variable_id",
    "std_variable_name_cn",
    "std_variable_name_en",
    "semantic_folder",
    "definition",
    "value_type",
    "standard_unit",
    "record_grain",
    "default_anchor_type",
    "active_databases",
    "current_status",
    "materialization_status",
    "latest_process_batch_id",
    "current_row_count",
    "layer3_asset_path",
    "knowledge_package_path",
    "log_archive_path",
    "owner",
    "latest_version",
    "latest_review_date",
    "remarks",
]

LOCALIZATION_HEADERS = [
    "std_variable_id",
    "language_code",
    "localized_name",
    "translation_status",
    "review_status",
    "source_field",
    "translator",
    "updated_at",
    "remarks",
]

VARIABLE_PROFILE_HEADERS = [
    "std_variable_id",
    "std_variable_name_cn",
    "std_variable_name_en",
    "semantic_folder",
    "definition",
    "value_type",
    "standard_unit",
    "record_grain",
    "default_anchor_type",
    "layer3_asset_path",
    "layer5_manifest_path",
    "layer5_preview_path",
    "current_status",
    "materialization_status",
    "latest_process_batch_id",
    "current_row_count",
    "owner",
    "current_version",
    "summary_note",
]

LAYER3_SCHEMA_HEADERS = [
    "column_order",
    "column_name",
    "column_name_cn",
    "data_type",
    "nullable_flag",
    "definition",
    "unit",
    "precision_rule",
    "key_role",
    "allowed_values_or_range",
    "usage_caution",
]

SEMANTIC_FOLDER_ROWS = [
    {
        "semantic_folder": "id_mapping",
        "display_name_cn": "ID譏蟆・",
        "scope_note": "subject_id縲”adm_id縲《tay_id 遲画ｸ蠢・ｸｻ髞ｮ譏蟆・所蜈ｶ蜈ｳ髞ｮ譌ｶ髣ｴ遯怜哨縲・",
    },
    {
        "semantic_folder": "encounter_information",
        "display_name_cn": "菴城劼/蟆ｱ隸贋ｿ｡諱ｯ",
        "scope_note": "菴城劼縲∝ｰｱ隸翫∬ｽｬ遘代！CU stay 遲芽ｿ・ｨ倶ｿ｡諱ｯ縲・",
    },
    {
        "semantic_folder": "demographics",
        "display_name_cn": "莠ｺ蜿｣扈溯ｮ｡蟄ｦ菫｡諱ｯ",
        "scope_note": "蟷ｴ鮴・∵ｧ蛻ｫ縲∫ｧ肴酪遲我ｺｺ蜿｣扈溯ｮ｡菫｡諱ｯ縲・",
    },
    {
        "semantic_folder": "anthropometrics",
        "display_name_cn": "菴捺ｼ豬矩㍼菫｡諱ｯ",
        "scope_note": "霄ｫ鬮倥∽ｽ馴㍾縲。MI 遲臥嶌蟇ｹ遞ｳ螳夂噪菴捺ｼ豬矩㍼蜿倬㍼縲・",
    },
    {
        "semantic_folder": "vital_signs",
        "display_name_cn": "逕溷多菴灘ｾ∽ｿ｡諱ｯ",
        "scope_note": "蠢・紫縲∬｡蜴九∝他蜷ｸ縲∽ｽ捺ｸｩ縲ヾpO2 遲峨・",
    },
    {
        "semantic_folder": "laboratory",
        "display_name_cn": "譽譟･譽鬪御ｿ｡諱ｯ",
        "scope_note": "螳樣ｪ悟ｮ､譽鬪御ｸ主・莉匁｣譟･扈捺棡縲・",
    },
    {
        "semantic_folder": "diagnosis_history",
        "display_name_cn": "譌｢蠕逞・彰/譌｢蠕隸頑妙",
        "scope_note": "蜴・彰逍ｾ逞・∵里蠕逞・彰縲∵里蠕郛也∬ｯ頑妙縲・",
    },
    {
        "semantic_folder": "diagnosis_current",
        "display_name_cn": "譛ｬ谺｡菴城劼/蠖灘燕隸頑妙",
        "scope_note": "蠖灘燕菴城劼謌門ｽ灘燕莠倶ｻｶ逶ｸ蜈ｳ隸頑妙縲・",
    },
    {
        "semantic_folder": "orders",
        "display_name_cn": "蛹ｻ蝌ｱ菫｡諱ｯ",
        "scope_note": "蛹ｻ蝌ｱ縲∝ｼ遶玖ｮｰ蠖輔∵欠莉､邀ｻ襍・ｺｧ縲・",
    },
    {
        "semantic_folder": "medication",
        "display_name_cn": "逕ｨ闕ｯ菫｡諱ｯ",
        "scope_note": "螟・婿縲∫ｻ呵艮縲∬艮迚ｩ證ｴ髴ｲ縲・",
    },
    {
        "semantic_folder": "nursing_execution_or_documented_care",
        "display_name_cn": "謚､逅・鴬陦・隶ｰ蠖慕・謚､",
        "scope_note": "謚､逅・ｮ梧・迥ｶ諤√∫・謚､謇ｧ陦瑚ｮｰ蠖輔∝ｺ頑浴謚､逅・枚譯｣遲峨・",
    },
    {
        "semantic_folder": "treatment_intervention",
        "display_name_cn": "豐ｻ逍・謫堺ｽ應ｿ｡諱ｯ",
        "scope_note": "謫堺ｽ懊∝ｹｲ鬚・∵ｲｻ逍苓｡御ｸｺ縲・",
    },
    {
        "semantic_folder": "device_support",
        "display_name_cn": "隶ｾ螟・謾ｯ謖∽ｿ｡諱ｯ",
        "scope_note": "譛ｺ譴ｰ騾壽ｰ斐，RRT縲・CMO 遲芽ｮｾ螟・ｸ取髪謖√・",
    },
    {
        "semantic_folder": "scores",
        "display_name_cn": "隸・・菫｡諱ｯ",
        "scope_note": "SOFA縲、PS縲ヾAPS 遲芽ｯ・・縲・",
    },
    {
        "semantic_folder": "outcomes",
        "display_name_cn": "扈灘ｱ菫｡諱ｯ",
        "scope_note": "豁ｻ莠｡縲∝・髯｢縲∬ｽｬ蠖偵∝・蜈･髯｢遲臥ｻ灘ｱ縲・",
    },
]

STD_LAB_RESULT_ROW = {
    "std_variable_id": "std_lab_result",
    "std_variable_name_cn": "譽鬪檎ｻ捺棡莠倶ｻｶ蠎募ｺｧ",
    "std_variable_name_en": "lab result event base",
    "semantic_folder": "laboratory",
    "definition": (
        "Event-level laboratory standardization baseline used to register hospitalization attribution, "
        "ICU stay linkage, and ICU-relative time for source lab results; analyte-specific concept "
        "harmonization and unit standardization are intentionally deferred."
    ),
    "value_type": "mixed_numeric_or_text",
    "standard_unit": None,
    "record_grain": "event",
    "default_anchor_type": "icu_intime",
    "active_databases": "MIMIC-IV-3.1",
    "current_status": "baseline_validated",
    "materialization_status": "not_retained_large_unsplit",
    "latest_process_batch_id": "20260322T000000Z_MIMIC-IV-3.1_std_lab_result",
    "current_row_count": 158374764,
    "layer3_asset_path": None,
    "knowledge_package_path": None,
    "log_archive_path": None,
    "owner": None,
    "latest_version": "v1",
    "latest_review_date": "2026-03-22",
    "remarks": (
        "A full-table Layer 3 materialization was built and validated, then intentionally deleted "
        "under the framework rule that very large unsplit standardized tables are not retained as "
        "formal Layer 3 assets. Keep the rule/log/index record in Layer 5 and build smaller analyte-"
        "specific retained variables later."
    ),
}


def repo_relative(path: Path) -> str:
    return path.relative_to(REPO_ROOT).as_posix()


def read_records(ws) -> list[dict[str, object]]:
    headers = [cell.value for cell in ws[1]]
    records: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if all(value is None for value in row):
            continue
        record = {}
        for index, header in enumerate(headers):
            if header is None:
                continue
            record[header] = row[index] if index < len(row) else None
        records.append(record)
    return records


def rewrite_sheet(ws, headers: list[str], records: list[dict[str, object]]) -> None:
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def ensure_sheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook.create_sheet(title=sheet_name)


def build_localization_records(
    master_records: list[dict[str, object]],
    existing_records: list[dict[str, object]],
    bootstrap_zh_cn: bool,
) -> list[dict[str, object]]:
    merged: dict[tuple[str, str], dict[str, object]] = {}
    today = date.today().isoformat()

    for record in existing_records:
        std_variable_id = record.get("std_variable_id")
        language_code = record.get("language_code")
        if not std_variable_id or not language_code:
            continue
        merged[(str(std_variable_id), str(language_code))] = dict(record)

    if bootstrap_zh_cn:
        for record in master_records:
            std_variable_id = record.get("std_variable_id")
            localized_name = record.get("std_variable_name_cn")
            if not std_variable_id or not localized_name:
                continue

            key = (str(std_variable_id), "zh-CN")
            existing = dict(merged.get(key, {}))
            previous_name = existing.get("localized_name")

            existing["std_variable_id"] = std_variable_id
            existing["language_code"] = "zh-CN"
            existing["localized_name"] = localized_name
            existing["translation_status"] = existing.get("translation_status") or "canonical_master"
            existing["review_status"] = existing.get("review_status") or "approved"
            existing["source_field"] = "std_variable_catalog.std_variable_name_cn"
            existing["translator"] = existing.get("translator") or "system_bootstrap"
            if not existing.get("updated_at") or previous_name != localized_name:
                existing["updated_at"] = today
            existing["remarks"] = existing.get("remarks")
            merged[key] = existing

    return sorted(
        merged.values(),
        key=lambda record: (
            str(record.get("std_variable_id") or ""),
            str(record.get("language_code") or ""),
        ),
    )


def ensure_master_index(path: Path, add_std_lab_result: bool) -> None:
    wb = load_workbook(path)
    ws = wb["std_variable_catalog"]
    records = read_records(ws)

    if add_std_lab_result:
        seen = False
        updated_records = []
        for record in records:
            if record.get("std_variable_id") == "std_lab_result":
                merged = dict(STD_LAB_RESULT_ROW)
                merged.update(record)
                updated_records.append(merged)
                seen = True
            else:
                updated_records.append(record)
        if not seen:
            updated_records.append(dict(STD_LAB_RESULT_ROW))
        records = updated_records

    rewrite_sheet(ws, MASTER_HEADERS, records)
    dict_ws = ensure_sheet(wb, "semantic_folder_dictionary")
    existing_semantic_records = read_records(dict_ws) if dict_ws.max_row > 1 else []
    rewrite_sheet(
        dict_ws,
        ["semantic_folder", "display_name_cn", "scope_note"],
        existing_semantic_records or SEMANTIC_FOLDER_ROWS,
    )
    localization_ws = ensure_sheet(wb, "std_variable_localization")
    existing_localization_records = read_records(localization_ws) if localization_ws.max_row > 1 else []
    rewrite_sheet(
        localization_ws,
        LOCALIZATION_HEADERS,
        build_localization_records(
            master_records=records,
            existing_records=existing_localization_records,
            bootstrap_zh_cn=add_std_lab_result,
        ),
    )
    wb.save(path)
    print(f"Updated master index workbook: {repo_relative(path)}")


def ensure_per_variable_template(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb["variable_profile"]
    records = read_records(ws)
    rewrite_sheet(ws, VARIABLE_PROFILE_HEADERS, records)
    contract_ws = ensure_sheet(wb, "standard_contract")
    rewrite_sheet(
        contract_ws,
        [
            "rule_order",
            "rule_dimension",
            "target_field",
            "canonical_rule",
            "allowed_values_or_range",
            "source_mapping_policy",
            "applies_to_scope",
            "if_unmappable_then",
            "change_requires_reprocessing_assessment",
            "approved_by",
            "approved_at",
            "remarks",
        ],
        [],
    )
    schema_ws = ensure_sheet(wb, "layer3_schema")
    rewrite_sheet(schema_ws, LAYER3_SCHEMA_HEADERS, [])
    wb.save(path)
    print(f"Updated per-variable template: {repo_relative(path)}")


def main() -> None:
    for path in MASTER_WORKBOOKS:
        ensure_master_index(path, add_std_lab_result=("Global" in str(path)))
    ensure_per_variable_template(PER_VARIABLE_TEMPLATE)


if __name__ == "__main__":
    main()

