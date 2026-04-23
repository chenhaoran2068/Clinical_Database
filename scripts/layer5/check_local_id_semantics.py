from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APPROVED_VALUES = {"approved", "reviewed_approved"}


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _workspace_root() -> Path:
    return _repo_root().parent.parent


def _default_database_root(workspace_root: Path, database_id: str) -> Path:
    return workspace_root / "Methods" / "Clinical_Database" / "local_work" / "Layer 5" / database_id


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_slug(value: Any) -> str:
    return _normalize_text(value).lower().replace("-", "_").replace(" ", "_")


def _parse_manifest_values(manifest_path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not manifest_path.exists():
        return values

    for line in manifest_path.read_text(encoding="utf-8", errors="replace").splitlines():
        if not line.startswith("- `") or "`: `" not in line:
            continue
        key = line.split("`", 2)[1]
        rest = line.split("`: `", 1)[1]
        values[key] = rest.rsplit("`", 1)[0]
    return values


def _sheet_rows_as_dicts(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(value in (None, "") for value in row):
                continue
            rows.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
        return rows
    finally:
        wb.close()


def _amsterdam_stay_equivalent_findings(
    *,
    asset_dir: Path,
    manifest_values: dict[str, str],
    schema_rows: list[dict[str, Any]],
) -> list[dict[str, str]]:
    findings: list[dict[str, str]] = []
    record_grain = _normalize_slug(manifest_values.get("record_grain"))

    for row in schema_rows:
        column_name = _normalize_slug(row.get("column_name"))
        key_role = _normalize_slug(row.get("key_role"))
        definition = _normalize_text(row.get("definition"))
        definition_lower = definition.lower()

        if column_name != "admissionid" or key_role not in {"stay_id", "stay_key"}:
            continue

        if "stay-equivalent" not in definition_lower:
            findings.append(
                {
                    "severity": "warning",
                    "kind": "definition_missing_stay_equivalent",
                    "variable": asset_dir.name,
                    "column_name": _normalize_text(row.get("column_name")),
                    "column_order": _normalize_text(row.get("column_order")),
                    "message": (
                        "Amsterdam `admissionid` is assigned a stay-level key role but the local schema definition "
                        "does not explicitly say `stay-equivalent`."
                    ),
                }
            )

        if "admission identifier" in definition_lower:
            findings.append(
                {
                    "severity": "warning",
                    "kind": "legacy_admission_identifier_phrase",
                    "variable": asset_dir.name,
                    "column_name": _normalize_text(row.get("column_name")),
                    "column_order": _normalize_text(row.get("column_order")),
                    "message": (
                        "Amsterdam `admissionid` is assigned a stay-level key role but the local schema definition "
                        "still uses the legacy phrase `admission identifier`."
                    ),
                }
            )

        if record_grain == "admission":
            findings.append(
                {
                    "severity": "warning",
                    "kind": "manifest_record_grain_conflicts_with_stay_role",
                    "variable": asset_dir.name,
                    "column_name": _normalize_text(row.get("column_name")),
                    "column_order": _normalize_text(row.get("column_order")),
                    "message": (
                        "The asset manifest still reports `record_grain: admission` while local `admissionid` is "
                        "being standardized as a stay-level key."
                    ),
                }
            )

    return findings


def _check_asset(
    *,
    database_id: str,
    asset_dir: Path,
    manifest_values: dict[str, str],
    schema_rows: list[dict[str, Any]],
) -> list[dict[str, str]]:
    if database_id.startswith("AmsterdamUMCdb"):
        return _amsterdam_stay_equivalent_findings(
            asset_dir=asset_dir,
            manifest_values=manifest_values,
            schema_rows=schema_rows,
        )
    return []


def run_check(
    *,
    database_id: str,
    database_root: Path,
    include_non_approved: bool,
) -> dict[str, Any]:
    if not database_root.exists():
        raise SystemExit(f"database root not found: {database_root}")

    asset_dirs = sorted(
        path for path in database_root.iterdir() if path.is_dir() and path.name.startswith("std_")
    )

    findings: list[dict[str, str]] = []
    checked_assets: list[str] = []
    skipped_assets: list[str] = []

    for asset_dir in asset_dirs:
        manifest_path = asset_dir / "asset_manifest.md"
        workbook_path = asset_dir / "Layer5_PerVariable_KnowledgePackage.xlsx"
        if not manifest_path.exists() or not workbook_path.exists():
            continue

        manifest_values = _parse_manifest_values(manifest_path)
        status = _normalize_slug(manifest_values.get("current_status"))
        if not include_non_approved and status not in APPROVED_VALUES:
            skipped_assets.append(asset_dir.name)
            continue

        schema_rows = _sheet_rows_as_dicts(workbook_path, "layer3_schema")
        checked_assets.append(asset_dir.name)
        findings.extend(
            _check_asset(
                database_id=database_id,
                asset_dir=asset_dir,
                manifest_values=manifest_values,
                schema_rows=schema_rows,
            )
        )

    return {
        "database_id": database_id,
        "database_root": str(database_root),
        "asset_dirs_seen": len(asset_dirs),
        "checked_assets": checked_assets,
        "skipped_assets": skipped_assets,
        "findings": findings,
    }


def _print_summary(result: dict[str, Any]) -> None:
    print("Clinical_Database local Layer 5 ID-semantics check")
    print(f"Database: {result['database_id']}")
    print(f"Database root: {result['database_root']}")
    print(f"Asset directories seen: {result['asset_dirs_seen']}")
    print(f"Checked assets: {len(result['checked_assets'])}")
    print(f"Skipped non-approved assets: {len(result['skipped_assets'])}")

    if result["checked_assets"]:
        print("Checked asset ids:")
        for asset_id in result["checked_assets"]:
            print(f"- {asset_id}")

    if result["skipped_assets"]:
        print("Skipped asset ids:")
        for asset_id in result["skipped_assets"]:
            print(f"- {asset_id}")

    findings = result["findings"]
    print(f"Findings: {len(findings)}")

    if not findings:
        print("Result: no local ID-semantics findings were detected for the active rule set.")
        if not result["database_id"].startswith("AmsterdamUMCdb"):
            print("Note: the current built-in rule set is Amsterdam-focused.")
        return

    print("Finding details:")
    for finding in findings:
        print(
            "- [{severity}] {kind} variable={variable} column={column_name} row={column_order}: {message}".format(
                **finding
            )
        )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Check local Layer 5 metadata for ID-normalization wording drift."
    )
    parser.add_argument("--database-id", required=True, help="Database id to inspect.")
    parser.add_argument(
        "--workspace-root",
        default=str(_workspace_root()),
        help="Workspace root containing Methods/Clinical_Database/local_work/Layer 5.",
    )
    parser.add_argument(
        "--database-root",
        help="Optional direct path to the local Layer 5 database root. Overrides --workspace-root + --database-id.",
    )
    parser.add_argument(
        "--include-non-approved",
        action="store_true",
        help="Also inspect non-approved assets instead of restricting the check to reviewed-approved assets.",
    )
    parser.add_argument(
        "--warning-only",
        action="store_true",
        help="Always exit 0 even when findings are present.",
    )
    args = parser.parse_args()

    workspace_root = Path(args.workspace_root).resolve()
    database_root = (
        Path(args.database_root).resolve()
        if args.database_root
        else _default_database_root(workspace_root, args.database_id)
    )

    result = run_check(
        database_id=args.database_id,
        database_root=database_root,
        include_non_approved=args.include_non_approved,
    )
    _print_summary(result)

    if result["findings"] and not args.warning_only:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
