from __future__ import annotations

import re
import time
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
import requests


REPO_ROOT = Path(__file__).resolve().parents[4]
WORKBOOK_PATH = REPO_ROOT / "Methods" / "Clinical_Database" / "local_work" / "Layer 5" / "Global" / "Layer5_StdVariable_MasterIndex.xlsx"
MASTER_SHEET = "std_variable_catalog"
LOCALIZATION_SHEET = "std_variable_localization"

LOCALIZATION_HEADERS = [
    "std_variable_id",
    "language_code",
    "localized_name",
    "translation_status",
    "review_status",
    "source_field",
    "translator",
    "updated_at",
    "remarks",
]

PROTECTED_REVIEW_STATUSES = {
    "approved",
    "professor_approved",
    "expert_reviewed",
    "locked",
}

REMARK = "Machine-generated Japanese draft from std_variable_name_cn; requires professor review."

POST_REPLACEMENTS = [
    ("繝・せ繝育ｵ先棡", "讀懈渊邨先棡"),
    ("インプットとアウトプット", "入出量"),
    ("アウトプット", "出量"),
    ("インプット", "入量"),
    ("使用状態", "使用状況"),
    ("评估状态", "評価状態"),
    ("縺ｾ縺ｨ繧・", "繧ｵ繝槭Μ繝ｼ"),
    ("閾ｨ蠎願ｩｦ鬨薙せ繧ｿ繧､繝ｫ", "閾ｨ蠎願ｩｦ鬨馴｢ｨ"),
    ("阮ｬ迚ｩ", "阮ｬ蜑､"),
    ("繧､繝吶Φ繝医・繝ｼ繧ｹ", "繧､繝吶Φ繝亥渕逶､"),
    ("28 譌･", "28譌･"),
    ("30 譌･", "30譌･"),
    ("48 譎る俣", "48譎る俣"),
    ("90 譌･", "90譌･"),
    ("365 譌･", "365譌･"),
    ("24 譎る俣", "24譎る俣"),
]


def read_records(ws) -> list[dict[str, object]]:
    headers = [cell.value for cell in ws[1]]
    records: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if all(value is None for value in row):
            continue
        record = {}
        for index, header in enumerate(headers):
            if header is None:
                continue
            record[header] = row[index] if index < len(row) else None
        records.append(record)
    return records


def rewrite_sheet(ws, headers: list[str], records: list[dict[str, object]]) -> None:
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def normalize_japanese(text: str) -> str:
    value = text.strip()
    for old, new in POST_REPLACEMENTS:
        value = value.replace(old, new)
    value = re.sub(r"\s+([/%・茨ｼ噂)])", r"\1", value)
    value = re.sub(r"([・・(])\s+", r"\1", value)
    value = re.sub(r"\s{2,}", " ", value)
    return value.strip()


def translate_batch(texts: list[str]) -> list[str]:
    translated: list[str] = []
    chunk_size = 80
    url = "https://translate.googleapis.com/translate_a/single"

    for start in range(0, len(texts), chunk_size):
        chunk = texts[start : start + chunk_size]
        try:
            payload = "\n".join(chunk)
            response = requests.get(
                url,
                params={
                    "client": "gtx",
                    "sl": "zh-CN",
                    "tl": "ja",
                    "dt": "t",
                    "q": payload,
                },
                timeout=60,
            )
            response.raise_for_status()
            data = response.json()
            merged = "".join(part[0] for part in data[0])
            results = merged.splitlines()
            if len(results) != len(chunk):
                raise RuntimeError("unexpected batch translation response")
        except Exception:
            results = []
            for item in chunk:
                response = requests.get(
                    url,
                    params={
                        "client": "gtx",
                        "sl": "zh-CN",
                        "tl": "ja",
                        "dt": "t",
                        "q": item,
                    },
                    timeout=60,
                )
                response.raise_for_status()
                data = response.json()
                results.append("".join(part[0] for part in data[0]))
                time.sleep(0.05)
        translated.extend(normalize_japanese(item) for item in results)
        time.sleep(0.05)

    return translated


def build_ja_jp_rows(
    master_records: list[dict[str, object]],
    localization_records: list[dict[str, object]],
) -> list[dict[str, object]]:
    today = date.today().isoformat()
    keyed_records: dict[tuple[str, str], dict[str, object]] = {}
    for record in localization_records:
        std_variable_id = str(record.get("std_variable_id") or "")
        language_code = str(record.get("language_code") or "")
        if std_variable_id and language_code:
            keyed_records[(std_variable_id, language_code)] = dict(record)

    to_translate_ids: list[str] = []
    to_translate_names: list[str] = []

    for record in master_records:
        std_variable_id = str(record.get("std_variable_id") or "")
        name_cn = str(record.get("std_variable_name_cn") or "").strip()
        if not std_variable_id or not name_cn:
            continue

        existing = keyed_records.get((std_variable_id, "ja-JP"))
        if existing:
            review_status = str(existing.get("review_status") or "").strip()
            if review_status in PROTECTED_REVIEW_STATUSES:
                continue

        to_translate_ids.append(std_variable_id)
        to_translate_names.append(name_cn)

    translated_names = translate_batch(to_translate_names)
    for std_variable_id, localized_name in zip(to_translate_ids, translated_names, strict=True):
        keyed_records[(std_variable_id, "ja-JP")] = {
            "std_variable_id": std_variable_id,
            "language_code": "ja-JP",
            "localized_name": localized_name,
            "translation_status": "machine_draft_google_zh_to_ja",
            "review_status": "pending_professor_review",
            "source_field": "std_variable_catalog.std_variable_name_cn",
            "translator": "google_translate_web_batch",
            "updated_at": today,
            "remarks": REMARK,
        }

    return sorted(
        keyed_records.values(),
        key=lambda record: (
            str(record.get("std_variable_id") or ""),
            str(record.get("language_code") or ""),
        ),
    )


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH)
    master_ws = workbook[MASTER_SHEET]
    localization_ws = workbook[LOCALIZATION_SHEET]

    master_records = read_records(master_ws)
    localization_records = read_records(localization_ws)
    updated_records = build_ja_jp_rows(master_records, localization_records)
    rewrite_sheet(localization_ws, LOCALIZATION_HEADERS, updated_records)
    workbook.save(WORKBOOK_PATH)

    total_ja = sum(1 for record in updated_records if record.get("language_code") == "ja-JP")
    print(f"Updated {LOCALIZATION_SHEET} with {total_ja} ja-JP rows in {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()

