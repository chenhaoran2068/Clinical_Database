from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from master_index_helper import default_master_index_path, read_database_asset_records, read_variable_catalog_records
from public_card_safety import contains_local_only_detail


LOCALIZATION_SHEET = "std_variable_localization"
ACTIVE_VALUES = {"y", "yes", "true", "1"}
APPROVED_VALUES = {"approved", "reviewed_approved"}


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _workspace_root() -> Path:
    return _repo_root().parent.parent


def _default_output_path(std_variable_id: str) -> Path:
    return _repo_root() / "docs" / "std_variable_cards" / f"{std_variable_id}.md"


def _default_output_dir() -> Path:
    return _repo_root() / "docs" / "std_variable_cards"


def _resolve_workspace_path(workspace_root: Path, raw_path: str | None) -> Path | None:
    if not raw_path:
        return None
    path = Path(str(raw_path))
    if path.is_absolute():
        return path
    return workspace_root / path


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _normalize_text(value).lower() in ACTIVE_VALUES


def _normalize_slug(value: Any) -> str:
    return _normalize_text(value).lower().replace("-", "_").replace(" ", "_")


def _strip_terminal_punctuation(text: str) -> str:
    return text.rstrip().rstrip(".")


def _sheet_rows_as_dicts(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(value in (None, "") for value in row):
                continue
            rows.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
        return rows
    finally:
        wb.close()


def _read_localization_records(workbook_path: Path, std_variable_id: str) -> list[dict[str, Any]]:
    rows = _sheet_rows_as_dicts(workbook_path, LOCALIZATION_SHEET)
    return [row for row in rows if row.get("std_variable_id") == std_variable_id]


def _approved_localized_name(records: list[dict[str, Any]], language_code: str) -> str:
    for row in records:
        if row.get("language_code") != language_code:
            continue
        review_status = _normalize_text(row.get("review_status")).lower()
        translation_status = _normalize_text(row.get("translation_status")).lower()
        if review_status == "approved" or translation_status == "canonical_master":
            return _normalize_text(row.get("localized_name"))
    return ""


def _normalize_public_value_type(value: Any) -> str:
    text = _normalize_slug(value)
    mapping = {
        "numeric_measurement": "numeric measurement",
        "numeric": "numeric measurement",
        "numeric_score": "numeric score",
        "categorical": "categorical",
        "categorical_text": "categorical text",
        "categorical_result": "categorical result",
        "boolean": "boolean",
        "binary_outcome": "binary outcome",
        "binary_flag": "binary flag",
        "duration_measurement": "duration measurement",
        "duration": "duration",
    }
    return mapping.get(text, _normalize_text(value))


def _normalize_public_grain(value: Any) -> str:
    text = _normalize_slug(value)
    mapping = {
        "event": "event-level measurement",
        "event_level_measurement": "event-level measurement",
        "admission": "admission-level record",
        "hospital_admission": "hospital admission",
        "icu_admission": "ICU admission",
        "patient": "patient",
        "stay": "ICU stay",
        "icu_stay": "ICU stay",
        "stay_equivalent": "ICU stay-equivalent",
        "icu_stay_equivalent": "ICU stay-equivalent",
        "icu_semantic_stay_equivalent": "ICU stay-equivalent",
    }
    return mapping.get(text, _normalize_text(value))


def _cross_database_definition_grain_phrase(grain: str) -> str:
    if grain == "patient":
        return "retained at patient level"
    if grain == "ICU stay":
        return "retained at ICU-stay level"
    if grain == "ICU stay-equivalent":
        return "retained at ICU-stay-equivalent level"
    if grain == "stay-level aligned":
        return "retained at stay level under approved database-specific local key conventions"
    if grain == "hospital admission":
        return "retained at hospital-admission level"
    if grain == "ICU admission":
        return "retained at ICU-admission level"
    if "admission" in grain:
        return "retained at admission level"
    return f"retained as {grain}"


def _normalize_public_anchor(value: Any) -> str:
    text = _normalize_slug(value)
    mapping = {
        "icu_intime": "ICU admission anchor",
        "icu_admission_anchor": "ICU admission anchor",
        "hospital_admission_code_summary": "hospital admission summary anchor",
    }
    return mapping.get(text, _normalize_text(value))


def _normalize_public_range_note(value: Any, standard_unit: str) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    unit = _normalize_text(standard_unit)
    if not unit:
        return text
    lowered = text.lower()
    if lowered.endswith(unit.lower()):
        return text
    if any(char.isalpha() for char in text):
        return text
    return f"{text} {unit}"


def _normalize_public_name_en(catalog_row: dict[str, Any], variable_profile: dict[str, Any]) -> str:
    catalog_value = _normalize_text(catalog_row.get("std_variable_name_en"))
    if catalog_value:
        return catalog_value
    return _normalize_text(variable_profile.get("std_variable_name_en"))


def _is_public_safe_text(value: Any) -> bool:
    text = _normalize_text(value)
    return bool(text) and not contains_local_only_detail(text)


def _build_public_definition_fallback(
    *,
    name_en: str,
    value_type: str,
    record_grain: str,
    standard_unit: str,
) -> str:
    sentences: list[str] = []
    if name_en:
        sentences.append(f"Standardized {name_en}.")
    else:
        sentences.append("Standardized variable.")

    public_value_type = _normalize_public_value_type(value_type)
    public_grain = _normalize_public_grain(record_grain)
    if public_value_type:
        sentences.append(f"Value type: {public_value_type}.")
    if public_grain:
        sentences.append(f"Grain: {public_grain}.")
    if standard_unit:
        sentences.append(f"Standard unit: {standard_unit}.")
    return " ".join(sentences)


def _select_public_definition(catalog_row: dict[str, Any], database_summary: dict[str, Any]) -> str:
    catalog_definition = _normalize_text(catalog_row.get("definition"))
    if _is_public_safe_text(catalog_definition):
        return catalog_definition

    database_definition = _normalize_text(database_summary.get("definition"))
    if _is_public_safe_text(database_definition):
        return database_definition

    return _build_public_definition_fallback(
        name_en=_normalize_public_name_en(catalog_row, {"std_variable_name_en": database_summary["name_en"]}),
        value_type=database_summary["value_type"] or _normalize_text(catalog_row.get("value_type")),
        record_grain=database_summary["record_grain"] or _normalize_text(catalog_row.get("record_grain")),
        standard_unit=database_summary["standard_unit"] or _normalize_text(catalog_row.get("standard_unit")),
    )


def _public_safe_cautions(cautions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    public_rows: list[dict[str, Any]] = []
    for caution in cautions:
        fields_to_check = (
            caution.get("topic"),
            caution.get("description"),
            caution.get("preferred_handling"),
        )
        if any(contains_local_only_detail(value) for value in fields_to_check):
            continue
        public_rows.append(caution)
    return public_rows


def _append_public_caution_lines(lines: list[str], cautions: list[dict[str, Any]]) -> bool:
    wrote_any = False
    for caution in _public_safe_cautions(cautions):
        caution_level = _normalize_text(caution.get("caution_level")) or "unspecified"
        topic = _normalize_text(caution.get("topic")) or "general"
        description = _normalize_text(caution.get("description"))
        preferred_handling = _normalize_text(caution.get("preferred_handling"))
        line = f"- `{caution_level}` / `{topic}`: {description}"
        if preferred_handling:
            line += f" Preferred handling: {_strip_terminal_punctuation(preferred_handling)}."
        lines.append(line)
        wrote_any = True
    return wrote_any


def _primary_value_sort_key(std_variable_id: str, row: dict[str, Any]) -> tuple[int, int, str]:
    column_name = _normalize_slug(row.get("column_name"))
    key_role = _normalize_slug(row.get("key_role"))
    std_prefix = _normalize_slug(std_variable_id)

    if key_role == "value_cleaned":
        priority = 0
    elif column_name.startswith(std_prefix) and "cleaned" in column_name:
        priority = 1
    elif column_name.startswith(std_prefix) and "normalized" in column_name:
        priority = 2
    elif column_name == std_prefix or column_name.startswith(std_prefix):
        priority = 3
    elif key_role == "value_normalized":
        priority = 4
    elif key_role == "value" and not column_name.startswith("source_"):
        priority = 5
    elif key_role == "value":
        priority = 6
    else:
        priority = 99

    column_order = row.get("column_order")
    if column_order is None:
        column_order = 999999
    return priority, int(column_order), column_name


def _load_knowledge_package_summary(std_variable_id: str, workbook_path: Path) -> dict[str, Any]:
    variable_profile_rows = _sheet_rows_as_dicts(workbook_path, "variable_profile")
    layer3_schema_rows = _sheet_rows_as_dicts(workbook_path, "layer3_schema")
    caution_rows = _sheet_rows_as_dicts(workbook_path, "global_cautions")

    variable_profile = variable_profile_rows[0] if variable_profile_rows else {}
    active_cautions = [row for row in caution_rows if _normalize_bool(row.get("active_flag"))]

    primary_value_row = {}
    if layer3_schema_rows:
        sorted_rows = sorted(layer3_schema_rows, key=lambda row: _primary_value_sort_key(std_variable_id, row))
        for row in sorted_rows:
            if _primary_value_sort_key(std_variable_id, row)[0] < 99:
                primary_value_row = row
                break

    return {
        "variable_profile": variable_profile,
        "primary_value_row": primary_value_row,
        "active_cautions": active_cautions,
    }


def _format_default_display_rule(primary_value_row: dict[str, Any]) -> str:
    precision_rule = _normalize_text(primary_value_row.get("precision_rule"))
    data_type = _normalize_slug(primary_value_row.get("data_type"))

    if precision_rule:
        return precision_rule
    if data_type.startswith("int"):
        return "integer"
    if data_type.startswith("float") or data_type.startswith("double"):
        return "preserve stored precision"
    return "not yet separately codified"


def _common_nonempty(values: list[str]) -> str:
    normalized_values = [_normalize_text(value) for value in values if _normalize_text(value)]
    if not normalized_values:
        return ""
    first = normalized_values[0]
    if all(value == first for value in normalized_values):
        return first
    return ""


def _common_nonempty_after(values: list[Any], normalizer) -> str:
    normalized_values = [normalizer(value) for value in values if normalizer(value)]
    if not normalized_values:
        return ""
    first = normalized_values[0]
    if all(value == first for value in normalized_values):
        return first
    return ""


def _approved_assets_note(records: list[dict[str, Any]]) -> str:
    database_ids = [record.get("database_id") for record in records if record.get("database_id")]
    if not database_ids:
        return "No reviewed-approved database asset is currently registered."
    if len(database_ids) == 1:
        return f"Currently approved in {database_ids[0]} only."
    return "Currently approved in: " + ", ".join(database_ids) + "."


def _build_database_summary(
    asset_record: dict[str, Any],
    knowledge_summary: dict[str, Any],
) -> dict[str, Any]:
    variable_profile = knowledge_summary["variable_profile"]
    primary_value_row = knowledge_summary["primary_value_row"]
    active_cautions = knowledge_summary["active_cautions"]
    return {
        "database_id": _normalize_text(asset_record.get("database_id")),
        "definition": _normalize_text(variable_profile.get("definition")) or _normalize_text(asset_record.get("definition")),
        "name_en": _normalize_text(variable_profile.get("std_variable_name_en")) or _normalize_text(
            asset_record.get("std_variable_name_en")
        ),
        "semantic_folder": _normalize_text(variable_profile.get("semantic_folder")) or _normalize_text(
            asset_record.get("semantic_folder")
        ),
        "standard_unit": _normalize_text(variable_profile.get("standard_unit")) or _normalize_text(
            asset_record.get("standard_unit")
        ),
        "value_type": _normalize_text(variable_profile.get("value_type")) or _normalize_text(asset_record.get("value_type")),
        "record_grain": _normalize_text(variable_profile.get("record_grain")) or _normalize_text(
            asset_record.get("record_grain")
        ),
        "default_anchor_type": _normalize_text(variable_profile.get("default_anchor_type")) or _normalize_text(
            asset_record.get("default_anchor_type")
        ),
        "default_display_rule": _format_default_display_rule(primary_value_row),
        "primary_value_column": _normalize_text(primary_value_row.get("column_name")),
        "primary_value_range": _normalize_text(primary_value_row.get("allowed_values_or_range")),
        "summary_note": _normalize_text(variable_profile.get("summary_note")) or _normalize_text(asset_record.get("remarks")),
        "current_status": _normalize_text(asset_record.get("current_status")),
        "latest_version": _normalize_text(asset_record.get("latest_version")),
        "latest_review_date": _normalize_text(asset_record.get("latest_review_date")),
        "active_cautions": active_cautions,
    }


def _build_single_database_card_markdown(
    *,
    std_variable_id: str,
    catalog_row: dict[str, Any],
    localization_rows: list[dict[str, Any]],
    approved_asset_records: list[dict[str, Any]],
    database_summary: dict[str, Any],
    metadata_prefers_database_summary: bool,
) -> str:
    name_cn = _approved_localized_name(localization_rows, "zh-CN") or _normalize_text(catalog_row.get("std_variable_name_cn"))
    name_en = _normalize_public_name_en(catalog_row, {"std_variable_name_en": database_summary["name_en"]})
    name_ja = _approved_localized_name(localization_rows, "ja-JP")
    semantic_folder = database_summary["semantic_folder"] or _normalize_text(catalog_row.get("semantic_folder"))
    standard_unit = database_summary["standard_unit"] or _normalize_text(catalog_row.get("standard_unit"))
    value_type = _normalize_public_value_type(
        database_summary["value_type"] or _normalize_text(catalog_row.get("value_type"))
    )
    record_grain = _normalize_public_grain(
        database_summary["record_grain"] or _normalize_text(catalog_row.get("record_grain"))
    )
    default_anchor_type = _normalize_public_anchor(
        database_summary["default_anchor_type"] or _normalize_text(catalog_row.get("default_anchor_type"))
    )
    definition = _select_public_definition(catalog_row, database_summary)
    primary_value_range = _normalize_public_range_note(database_summary["primary_value_range"], standard_unit)
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    lines: list[str] = []
    lines.append(f"# Public Variable Card: {std_variable_id}")
    lines.append("")
    lines.append(f"Generated by `scripts/layer5/export_public_variable_card.py` on {generated_at}.")
    lines.append("This is a GitHub-safe public metadata summary, not the full local Layer 5 execution evidence package.")
    lines.append("")
    lines.append("## Identity")
    lines.append("")
    lines.append(f"- `std_variable_id`: `{std_variable_id}`")
    lines.append(f"- standardized Chinese name: `{name_cn}`" if name_cn else "- standardized Chinese name: not yet published")
    lines.append(f"- standardized English name: `{name_en}`" if name_en else "- standardized English name: not yet published")
    lines.append(f"- standardized Japanese name: `{name_ja}`" if name_ja else "- standardized Japanese name: not yet approved for public publication")
    lines.append(f"- semantic folder: `{semantic_folder}`" if semantic_folder else "- semantic folder: not yet published")
    lines.append(f"- standard unit: `{standard_unit}`" if standard_unit else "- standard unit: none")
    lines.append(f"- value type: `{value_type}`" if value_type else "- value type: not yet published")
    lines.append(f"- grain: `{record_grain}`" if record_grain else "- grain: not yet published")
    lines.append("")
    lines.append("## Standard Definition")
    lines.append("")
    lines.append(definition if definition else "Definition not yet published.")
    lines.append("")
    lines.append("## Default Presentation")
    lines.append("")
    lines.append(f"- default display rule: `{database_summary['default_display_rule']}`")
    if database_summary["primary_value_column"]:
        lines.append(f"- primary retained value column: `{database_summary['primary_value_column']}`")
    if default_anchor_type:
        lines.append(f"- default anchor type: `{default_anchor_type}`")
    if primary_value_range:
        lines.append(f"- current approved value range note: `{primary_value_range}`")
    lines.append("")
    lines.append("## Global Warnings")
    lines.append("")
    if not _append_public_caution_lines(lines, database_summary["active_cautions"]):
        lines.append("- No public-safe global cautions are currently published for this public card.")
    lines.append("")
    lines.append("## Cross-Database Status")
    lines.append("")
    lines.append(f"- {_approved_assets_note(approved_asset_records)}")
    lines.append(f"- current publication basis: reviewed-approved `{database_summary['database_id']}` knowledge package")
    if metadata_prefers_database_summary:
        lines.append(
            "- metadata source rule: this card prefers the selected database knowledge package over legacy `std_variable_catalog` text because multiple approved database assets currently coexist."
        )
    lines.append("")
    lines.append("## Current Approved Database Assets")
    lines.append("")
    lines.append("| database_id | current_status | latest_version | latest_review_date |")
    lines.append("| --- | --- | --- | --- |")
    for record in sorted(approved_asset_records, key=lambda row: _normalize_text(row.get("database_id"))):
        lines.append(
            "| {database_id} | {current_status} | {latest_version} | {latest_review_date} |".format(
                database_id=_normalize_text(record.get("database_id")) or "-",
                current_status=_normalize_text(record.get("current_status")) or "-",
                latest_version=_normalize_text(record.get("latest_version")) or "-",
                latest_review_date=_normalize_text(record.get("latest_review_date")) or "-",
            )
        )
    lines.append("")
    lines.append("## Publication Rule")
    lines.append("")
    lines.append(
        "Detailed source-table mappings, database-specific build logs, grouped review history, and rerun assessment remain in local Layer 5 evidence packages rather than this public card."
    )
    lines.append(
        "The public-card exporter applies a conservative publication filter and suppresses local-only implementation detail such as raw source fields, itemids, and local threshold mechanics."
    )
    lines.append("")
    return "\n".join(lines)


def _build_cross_database_definition(name_en: str, grain: str, unit: str) -> str:
    subject = name_en or "this standardized variable"
    if grain:
        sentence = f"Cross-database standardized {subject} {_cross_database_definition_grain_phrase(grain)}."
        if unit:
            sentence += f" Standard unit: {unit}."
        return sentence
    if unit:
        return (
            f"Cross-database standardized {subject} with one shared variable identity across the approved "
            f"database implementations listed below. Standard unit: {unit}."
        )
    return (
        f"Cross-database standardized {subject} with one shared variable identity across the approved "
        "database implementations listed below."
    )


def _build_cross_database_interpretation_rule(database_summaries: list[dict[str, Any]]) -> str:
    normalized_grains: list[str] = []
    for summary in database_summaries:
        grain = _normalize_public_grain(summary.get("record_grain"))
        if grain:
            normalized_grains.append(grain)

    unique_grains = list(dict.fromkeys(normalized_grains))
    if not unique_grains:
        return (
            "current public grain rule: respect each approved database implementation's published grain and do not "
            "silently reinterpret the variable at a different analytic level"
        )

    stay_aligned_grains = {"ICU stay", "ICU stay-equivalent"}
    if len(unique_grains) > 1 and set(unique_grains).issubset(stay_aligned_grains):
        joined = ", ".join(f"`{grain}`" for grain in unique_grains)
        return (
            "current public grain rule: approved database implementations are semantically stay-level aligned; "
            f"respect the published per-database grain ({joined}) and do not silently reinterpret local "
            "stay-equivalent keys as hospital-admission keys or event-level rows"
        )

    if len(unique_grains) > 1:
        joined = ", ".join(f"`{grain}`" for grain in unique_grains)
        return (
            "current public grain rule: approved database implementations do not yet expose one identical grain; "
            f"respect the published per-database grain ({joined}) and do not silently collapse them into one generic interpretation"
        )

    grain = unique_grains[0]
    if "event-level" in grain:
        return (
            "current public grain rule: this is an event-level asset and should not be silently reinterpreted as a "
            "baseline or summary variable"
        )
    if grain == "patient":
        return (
            "current public grain rule: this is a patient-level asset and should not be silently reinterpreted as an "
            "admission-level, stay-level, or event-level variable"
        )
    if "stay" in grain:
        return (
            "current public grain rule: this is a stay-level asset and should not be silently reinterpreted as a "
            "patient-level or event-level variable"
        )
    if "admission" in grain:
        return (
            "current public grain rule: this is an admission-level asset and should not be silently reinterpreted as "
            "a patient-level or event-level variable"
        )
    return (
        f"current public grain rule: this is a `{grain}` asset and should not be silently reinterpreted at a "
        "different analytic level"
    )


def _build_shared_cross_database_grain(database_summaries: list[dict[str, Any]]) -> str:
    normalized_grains: list[str] = []
    for summary in database_summaries:
        grain = _normalize_public_grain(summary.get("record_grain"))
        if grain:
            normalized_grains.append(grain)

    unique_grains = list(dict.fromkeys(normalized_grains))
    if not unique_grains:
        return ""
    if len(unique_grains) == 1:
        return unique_grains[0]

    stay_aligned_grains = {"ICU stay", "ICU stay-equivalent"}
    if set(unique_grains).issubset(stay_aligned_grains):
        return "stay-level aligned"

    return ""


def _build_cross_database_card_markdown(
    *,
    std_variable_id: str,
    catalog_row: dict[str, Any],
    localization_rows: list[dict[str, Any]],
    approved_asset_records: list[dict[str, Any]],
    database_summaries: list[dict[str, Any]],
) -> str:
    name_cn = _approved_localized_name(localization_rows, "zh-CN") or _normalize_text(catalog_row.get("std_variable_name_cn"))
    name_en = _normalize_text(catalog_row.get("std_variable_name_en"))
    name_ja = _approved_localized_name(localization_rows, "ja-JP")
    semantic_folder = _common_nonempty([summary["semantic_folder"] for summary in database_summaries])
    standard_unit = _common_nonempty([summary["standard_unit"] for summary in database_summaries])
    value_type = _common_nonempty_after([summary["value_type"] for summary in database_summaries], _normalize_public_value_type)
    record_grain = _build_shared_cross_database_grain(database_summaries)
    display_rule = _common_nonempty([summary["default_display_rule"] for summary in database_summaries])
    common_range = _common_nonempty_after(
        [summary["primary_value_range"] for summary in database_summaries],
        lambda value: _normalize_public_range_note(value, standard_unit),
    )
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    definition = _build_cross_database_definition(name_en, record_grain, standard_unit)
    interpretation_rule = _build_cross_database_interpretation_rule(database_summaries)

    lines: list[str] = []
    lines.append(f"# Public Variable Card: {std_variable_id}")
    lines.append("")
    lines.append(f"Generated by `scripts/layer5/export_public_variable_card.py` on {generated_at}.")
    lines.append("This is a GitHub-safe public metadata summary, not the full local Layer 5 execution evidence package.")
    lines.append("")
    lines.append("## Identity")
    lines.append("")
    lines.append(f"- `std_variable_id`: `{std_variable_id}`")
    lines.append(f"- standardized Chinese name: `{name_cn}`" if name_cn else "- standardized Chinese name: not yet published")
    lines.append(f"- standardized English name: `{name_en}`" if name_en else "- standardized English name: not yet published")
    lines.append(f"- standardized Japanese name: `{name_ja}`" if name_ja else "- standardized Japanese name: not yet approved for public publication")
    lines.append(f"- semantic folder: `{semantic_folder}`" if semantic_folder else "- semantic folder: current approved database assets do not yet expose one identical public value")
    lines.append(f"- standard unit: `{standard_unit}`" if standard_unit else "- standard unit: current approved database assets do not yet expose one identical public value")
    lines.append(f"- value type: `{value_type}`" if value_type else "- value type: current approved database assets still need explicit public normalization")
    lines.append(f"- grain: `{record_grain}`" if record_grain else "- grain: current approved database assets still need explicit public normalization")
    lines.append("")
    lines.append("## Cross-Database Standard Definition")
    lines.append("")
    lines.append(definition)
    lines.append("")
    lines.append("## Shared Current Contract")
    lines.append("")
    if display_rule:
        lines.append(f"- current reviewed-approved public display rule across listed databases: `{display_rule}`")
    else:
        lines.append("- current reviewed-approved public display rule is not yet identical across all listed databases")
    if common_range:
        lines.append(f"- current reviewed-approved cleaned-value range note across listed databases: `{common_range}`")
    else:
        lines.append("- current reviewed-approved cleaned-value range note is not yet identical across all listed databases")
    lines.append(f"- {interpretation_rule}")
    lines.append("- current public interpretation rule: database-specific timing and anchoring semantics still require each database's approved local Layer 5 evidence package")
    lines.append("")
    lines.append("## Approved Database Implementations")
    lines.append("")
    lines.append("| database_id | publication name | value type | grain | default anchor family | primary retained value column | display rule | range note | latest_review_date |")
    lines.append("| --- | --- | --- | --- | --- | --- | --- | --- | --- |")
    for summary in sorted(database_summaries, key=lambda row: row["database_id"]):
        lines.append(
            "| {database_id} | {name_en} | {value_type} | {record_grain} | {anchor} | {value_col} | {display_rule} | {range_note} | {review_date} |".format(
                database_id=summary["database_id"] or "-",
                name_en=summary["name_en"] or "-",
                value_type=_normalize_public_value_type(summary["value_type"]) or "-",
                record_grain=_normalize_public_grain(summary["record_grain"]) or "-",
                anchor=_normalize_public_anchor(summary["default_anchor_type"]) or "-",
                value_col=summary["primary_value_column"] or "-",
                display_rule=summary["default_display_rule"] or "-",
                range_note=_normalize_public_range_note(summary["primary_value_range"], summary["standard_unit"]) or "-",
                review_date=summary["latest_review_date"] or "-",
            )
        )
    lines.append("")
    lines.append("## Database-Specific Notes")
    lines.append("")
    for summary in sorted(database_summaries, key=lambda row: row["database_id"]):
        lines.append(f"### {summary['database_id']}")
        lines.append("")
        if not _append_public_caution_lines(lines, summary["active_cautions"]):
            lines.append("- no database-specific public caution is currently published for this database implementation")
        lines.append("")
    lines.append("## Cross-Database Status")
    lines.append("")
    lines.append(f"- {_approved_assets_note(approved_asset_records)}")
    lines.append("- current publication basis: cross-database public card synthesized from all reviewed-approved database assets listed below")
    lines.append("- metadata source rule: shared public fields are emitted only when current approved database assets agree after light public normalization")
    lines.append("")
    lines.append("## Current Approved Database Assets")
    lines.append("")
    lines.append("| database_id | current_status | latest_version | latest_review_date |")
    lines.append("| --- | --- | --- | --- |")
    for record in sorted(approved_asset_records, key=lambda row: _normalize_text(row.get("database_id"))):
        lines.append(
            "| {database_id} | {current_status} | {latest_version} | {latest_review_date} |".format(
                database_id=_normalize_text(record.get("database_id")) or "-",
                current_status=_normalize_text(record.get("current_status")) or "-",
                latest_version=_normalize_text(record.get("latest_version")) or "-",
                latest_review_date=_normalize_text(record.get("latest_review_date")) or "-",
            )
        )
    lines.append("")
    lines.append("## Publication Rule")
    lines.append("")
    lines.append(
        "Detailed source-table mappings, database-specific build logs, grouped review history, and rerun assessment remain in local Layer 5 evidence packages rather than this public card."
    )
    lines.append(
        "The public-card exporter applies a conservative publication filter and suppresses local-only implementation detail such as raw source fields, itemids, and local threshold mechanics."
    )
    lines.append("")
    return "\n".join(lines)


def build_public_variable_card(
    *,
    workspace_root: Path,
    std_variable_id: str,
    database_id: str | None,
    output_path: Path,
    overwrite: bool,
    cross_database: bool,
) -> Path:
    master_index_path = default_master_index_path(workspace_root)
    asset_records = read_database_asset_records(master_index_path, std_variable_id=std_variable_id)
    approved_asset_records = [
        record for record in asset_records if _normalize_text(record.get("current_status")).lower() in APPROVED_VALUES
    ]
    if not approved_asset_records:
        raise ValueError(f"No reviewed-approved database asset found for `{std_variable_id}`.")
    if cross_database and database_id:
        raise ValueError("Do not pass `--database-id` together with `--cross-database`.")

    catalog_rows = read_variable_catalog_records(master_index_path, std_variable_id=std_variable_id)
    catalog_row = catalog_rows[0] if catalog_rows else {}
    localization_rows = _read_localization_records(master_index_path, std_variable_id)

    if cross_database:
        database_summaries: list[dict[str, Any]] = []
        for asset_record in approved_asset_records:
            knowledge_package_path = _resolve_workspace_path(
                workspace_root, _normalize_text(asset_record.get("knowledge_package_path"))
            )
            if knowledge_package_path is None or not knowledge_package_path.exists():
                raise FileNotFoundError(
                    f"Knowledge package not found for `{std_variable_id}` at `{asset_record.get('knowledge_package_path')}`."
                )
            knowledge_summary = _load_knowledge_package_summary(std_variable_id, knowledge_package_path)
            database_summaries.append(_build_database_summary(asset_record, knowledge_summary))
        markdown = _build_cross_database_card_markdown(
            std_variable_id=std_variable_id,
            catalog_row=catalog_row,
            localization_rows=localization_rows,
            approved_asset_records=approved_asset_records,
            database_summaries=database_summaries,
        )
    else:
        if database_id:
            matching = [record for record in approved_asset_records if record.get("database_id") == database_id]
            if not matching:
                raise ValueError(
                    f"No reviewed-approved database asset found for `{std_variable_id}` in database `{database_id}`."
                )
            selected_asset_record = matching[0]
        else:
            if len(approved_asset_records) > 1:
                databases = ", ".join(sorted(_normalize_text(record.get("database_id")) for record in approved_asset_records))
                raise ValueError(
                    "Multiple reviewed-approved database assets exist for "
                    f"`{std_variable_id}` ({databases}). Pass `--database-id` to choose the current publication basis or `--cross-database` to synthesize a merged public card."
                )
            selected_asset_record = approved_asset_records[0]

        knowledge_package_path = _resolve_workspace_path(
            workspace_root, _normalize_text(selected_asset_record.get("knowledge_package_path"))
        )
        if knowledge_package_path is None or not knowledge_package_path.exists():
            raise FileNotFoundError(
                f"Knowledge package not found for `{std_variable_id}` at `{selected_asset_record.get('knowledge_package_path')}`."
            )

        knowledge_summary = _load_knowledge_package_summary(std_variable_id, knowledge_package_path)
        database_summary = _build_database_summary(selected_asset_record, knowledge_summary)
        markdown = _build_single_database_card_markdown(
            std_variable_id=std_variable_id,
            catalog_row=catalog_row,
            localization_rows=localization_rows,
            approved_asset_records=approved_asset_records,
            database_summary=database_summary,
            metadata_prefers_database_summary=len(approved_asset_records) > 1,
        )

    output_path = Path(output_path)
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"Output already exists: {output_path}. Pass `--overwrite` to replace it.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(markdown, encoding="utf-8")
    return output_path


def batch_export_public_variable_cards(
    *,
    workspace_root: Path,
    output_dir: Path,
    overwrite: bool,
    only_missing: bool,
    database_id: str | None,
    cross_database: bool,
) -> dict[str, Any]:
    master_index_path = default_master_index_path(workspace_root)
    asset_records = read_database_asset_records(master_index_path)
    approved_asset_records = [
        record for record in asset_records if _normalize_text(record.get("current_status")).lower() in APPROVED_VALUES
    ]

    grouped_records: dict[str, list[dict[str, Any]]] = {}
    for record in approved_asset_records:
        std_variable_id = _normalize_text(record.get("std_variable_id"))
        if not std_variable_id:
            continue
        grouped_records.setdefault(std_variable_id, []).append(record)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    written: list[str] = []
    skipped: list[dict[str, str]] = []

    for std_variable_id in sorted(grouped_records):
        records = grouped_records[std_variable_id]
        output_path = output_dir / f"{std_variable_id}.md"

        if only_missing and output_path.exists():
            skipped.append(
                {
                    "std_variable_id": std_variable_id,
                    "reason": "output_exists_only_missing",
                }
            )
            continue

        if len(records) > 1:
            if cross_database:
                selected_database_id = None
                selected_cross_database = True
            elif database_id:
                if any(_normalize_text(record.get("database_id")) == database_id for record in records):
                    selected_database_id = database_id
                    selected_cross_database = False
                else:
                    skipped.append(
                        {
                            "std_variable_id": std_variable_id,
                            "reason": f"multi_database_no_requested_database_match:{database_id}",
                        }
                    )
                    continue
            else:
                skipped.append(
                    {
                        "std_variable_id": std_variable_id,
                        "reason": "multi_database_requires_database_id_or_cross_database",
                    }
                )
                continue
        else:
            selected_database_id = None
            selected_cross_database = False

        try:
            written_path = build_public_variable_card(
                workspace_root=workspace_root,
                std_variable_id=std_variable_id,
                database_id=selected_database_id,
                output_path=output_path,
                overwrite=overwrite,
                cross_database=selected_cross_database,
            )
        except FileExistsError:
            skipped.append(
                {
                    "std_variable_id": std_variable_id,
                    "reason": "output_exists_use_overwrite",
                }
            )
            continue
        written.append(str(written_path))

    return {
        "workspace_root": str(workspace_root),
        "output_dir": str(output_dir),
        "approved_asset_rows": len(approved_asset_records),
        "unique_std_variables": len(grouped_records),
        "written_count": len(written),
        "skipped_count": len(skipped),
        "written_paths": written,
        "skipped": skipped,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Export a GitHub-safe public variable card from local Layer 5 metadata.")
    parser.add_argument("--workspace-root", default=str(_workspace_root()), help="Workspace root containing Methods/...")
    selection_group = parser.add_mutually_exclusive_group(required=True)
    selection_group.add_argument("--std-variable-id", help="Standardized variable id to publish.")
    selection_group.add_argument(
        "--all-reviewed-approved",
        action="store_true",
        help="Batch-export all std_variable_id entries that currently have at least one reviewed-approved database asset.",
    )
    parser.add_argument(
        "--database-id",
        help="Database id whose reviewed-approved knowledge package should be used as the current publication basis.",
    )
    parser.add_argument(
        "--cross-database",
        action="store_true",
        help="When multiple reviewed-approved database assets exist, synthesize one merged public card from all of them. In batch mode, this rule is applied variable-by-variable for multi-database cases.",
    )
    parser.add_argument(
        "--output",
        help="Output markdown path. Default: docs/std_variable_cards/<std_variable_id>.md in the public repository.",
    )
    parser.add_argument(
        "--output-dir",
        help="Batch output directory. Default: docs/std_variable_cards in the public repository.",
    )
    parser.add_argument(
        "--only-missing",
        action="store_true",
        help="In batch mode, skip variables whose output markdown already exists.",
    )
    parser.add_argument("--overwrite", action="store_true", help="Overwrite the output file if it already exists.")
    args = parser.parse_args()

    workspace_root = Path(args.workspace_root).resolve()
    if args.all_reviewed_approved:
        if args.output:
            raise ValueError("Do not pass --output in batch mode. Use --output-dir instead.")
        output_dir = Path(args.output_dir).resolve() if args.output_dir else _default_output_dir()
        summary = batch_export_public_variable_cards(
            workspace_root=workspace_root,
            output_dir=output_dir,
            overwrite=args.overwrite,
            only_missing=args.only_missing,
            database_id=args.database_id,
            cross_database=args.cross_database,
        )
        print(
            "Batch export completed: "
            f"written={summary['written_count']}, skipped={summary['skipped_count']}, "
            f"unique_std_variables={summary['unique_std_variables']}, output_dir={summary['output_dir']}"
        )
        if summary["skipped"]:
            print("Skipped variables:")
            for skipped_item in summary["skipped"]:
                print(f"- {skipped_item['std_variable_id']}: {skipped_item['reason']}")
        return

    if args.only_missing:
        raise ValueError("--only-missing is supported only together with --all-reviewed-approved.")
    if args.output_dir:
        raise ValueError("--output-dir is supported only together with --all-reviewed-approved.")

    output_path = Path(args.output).resolve() if args.output else _default_output_path(args.std_variable_id)
    written = build_public_variable_card(
        workspace_root=workspace_root,
        std_variable_id=args.std_variable_id,
        database_id=args.database_id,
        output_path=output_path,
        overwrite=args.overwrite,
        cross_database=args.cross_database,
    )
    print(f"Wrote public variable card: {written}")


if __name__ == "__main__":
    main()
