from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
import subprocess
import tempfile
from typing import Any

from openpyxl import Workbook, load_workbook


MASTER_INDEX_REL = Path("Methods") / "Clinical_Database" / "local_work" / "Layer 5" / "Global" / "Layer5_StdVariable_MasterIndex.xlsx"
DATABASE_ASSET_SHEET = "std_variable_database_assets"
VARIABLE_CATALOG_SHEET = "std_variable_catalog"

DATABASE_ASSET_HEADERS = [
    "std_variable_id",
    "database_id",
    "std_variable_name_cn",
    "std_variable_name_en",
    "semantic_folder",
    "definition",
    "value_type",
    "standard_unit",
    "record_grain",
    "default_anchor_type",
    "current_status",
    "materialization_status",
    "latest_process_batch_id",
    "current_row_count",
    "layer3_asset_path",
    "knowledge_package_path",
    "layer5_asset_manifest_path",
    "log_archive_path",
    "owner",
    "latest_version",
    "latest_review_date",
    "remarks",
]

REQUIRED_DATABASE_ASSET_KEYS = {"std_variable_id", "database_id"}


def default_master_index_path(workspace_root: str | Path) -> Path:
    return Path(workspace_root).resolve() / MASTER_INDEX_REL


def _normalize_scalar(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, Path):
        return value.as_posix()
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return value


def _normalize_date_like_text(value: Any) -> Any:
    if value in (None, ""):
        return value
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        try:
            text = value.isoformat()
        except Exception:
            text = str(value)
    else:
        text = str(value)
    text = text.strip()
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]
    return text


def _sheet_headers(ws) -> list[str]:
    return [cell.value for cell in ws[1]]


def _prefer_excel_com_backend() -> bool:
    # Default to openpyxl for cross-platform reproducibility and to avoid
    # Windows Excel COM Unicode corruption in multilingual metadata fields.
    backend = os.getenv("CLINICAL_DATABASE_MASTER_INDEX_BACKEND", "openpyxl").strip().lower()
    if backend == "excel_com":
        return True
    return False


def _ensure_workbook(workbook_path: Path):
    if workbook_path.exists():
        return load_workbook(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    default.title = "README"
    default["A1"] = "Auto-created by master_index_helper.py"
    return wb


def _get_or_create_sheet(wb, sheet_name: str, after_sheet_name: str | None = None):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    if after_sheet_name and after_sheet_name in wb.sheetnames:
        after_ws = wb[after_sheet_name]
        ws = wb.create_sheet(sheet_name, wb.index(after_ws) + 1)
    else:
        ws = wb.create_sheet(sheet_name)
    return ws


def ensure_database_asset_sheet(workbook_path: str | Path):
    workbook_path = Path(workbook_path)
    if workbook_path.exists() and _prefer_excel_com_backend():
        try:
            _ensure_database_asset_sheet_excel_com(workbook_path)
            return
        except Exception:
            pass

    _ensure_database_asset_sheet_openpyxl(workbook_path)


def _ensure_database_asset_sheet_openpyxl(workbook_path: Path):
    workbook_path = Path(workbook_path)
    wb = _ensure_workbook(workbook_path)
    try:
        ws = _get_or_create_sheet(wb, DATABASE_ASSET_SHEET, VARIABLE_CATALOG_SHEET)
        headers = _sheet_headers(ws) if ws.max_row >= 1 else []
        if not any(headers):
            for col_index, header in enumerate(DATABASE_ASSET_HEADERS, start=1):
                ws.cell(row=1, column=col_index, value=header)
        else:
            current_headers = _sheet_headers(ws)[: len(DATABASE_ASSET_HEADERS)]
            if current_headers != DATABASE_ASSET_HEADERS:
                raise ValueError(
                    f"{DATABASE_ASSET_SHEET} header mismatch. "
                    "Refuse to auto-write because the existing schema is not the expected V2 schema."
                )
        wb.save(workbook_path)
    finally:
        wb.close()


def _run_excel_com_sheet_operation(
    workbook_path: Path,
    operation: str,
    record: dict[str, Any] | None = None,
) -> dict[str, Any]:
    payload = {
        "operation": operation,
        "workbook_path": str(workbook_path),
        "sheet_name": DATABASE_ASSET_SHEET,
        "after_sheet_name": VARIABLE_CATALOG_SHEET,
        "headers": DATABASE_ASSET_HEADERS,
        "record": _json_safe(record or {}),
    }
    script = r"""
$payload = Get-Content -LiteralPath $env:MASTER_INDEX_HELPER_PAYLOAD -Raw | ConvertFrom-Json
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {
    if (-not (Test-Path -LiteralPath $payload.workbook_path)) {
        throw "Workbook not found: $($payload.workbook_path)"
    }

    $wb = $excel.Workbooks.Open($payload.workbook_path)
    try {
        $sheet = $null
        foreach ($ws in $wb.Worksheets) {
            if ($ws.Name -eq $payload.sheet_name) {
                $sheet = $ws
                break
            }
        }
        if ($null -eq $sheet) {
            $after = $null
            foreach ($ws in $wb.Worksheets) {
                if ($ws.Name -eq $payload.after_sheet_name) {
                    $after = $ws
                    break
                }
            }
            if ($null -ne $after) {
                $sheet = $wb.Worksheets.Add([System.Type]::Missing, $after)
            } else {
                $sheet = $wb.Worksheets.Add()
            }
            $sheet.Name = $payload.sheet_name
        }

        $headers = @($payload.headers)
        $headerMismatch = $false
        $hasAnyHeader = $false
        for ($i = 0; $i -lt $headers.Count; $i++) {
            $cellText = [string]$sheet.Cells.Item(1, $i + 1).Text
            if (-not [string]::IsNullOrWhiteSpace($cellText)) {
                $hasAnyHeader = $true
            }
            if ($hasAnyHeader -and $cellText -ne $headers[$i]) {
                $headerMismatch = $true
            }
        }
        if ($headerMismatch) {
            throw "Existing sheet header does not match expected V2 schema."
        }
        if (-not $hasAnyHeader) {
            for ($i = 0; $i -lt $headers.Count; $i++) {
                $sheet.Cells.Item(1, $i + 1).Value2 = $headers[$i]
            }
        }

        $result = @{
            action = "ensured"
            row_index = $null
            std_variable_id = $null
            database_id = $null
            workbook_path = $payload.workbook_path
        }

        if ($payload.operation -eq 'upsert') {
            $record = $payload.record
            $targetRow = $null
            $lastRow = $sheet.UsedRange.Rows.Count
            if ($lastRow -lt 2) { $lastRow = 1 }
            for ($row = 2; $row -le $lastRow; $row++) {
                $stdId = [string]$sheet.Cells.Item($row, 1).Text
                $dbId = [string]$sheet.Cells.Item($row, 2).Text
                if ($stdId -eq [string]$record.std_variable_id -and $dbId -eq [string]$record.database_id) {
                    $targetRow = $row
                    break
                }
            }
            if ($null -eq $targetRow) {
                $targetRow = $lastRow + 1
                if ($targetRow -lt 2) { $targetRow = 2 }
                $result.action = "inserted"
            } else {
                $result.action = "updated"
            }

            for ($i = 0; $i -lt $headers.Count; $i++) {
                $header = $headers[$i]
                $value = $record.$header
                if ($null -eq $value) { $value = "" }
                $sheet.Cells.Item($targetRow, $i + 1).Value2 = [string]$value
            }

            $result.row_index = $targetRow
            $result.std_variable_id = [string]$record.std_variable_id
            $result.database_id = [string]$record.database_id
        }

        $sheet.Columns.AutoFit() | Out-Null
        $wb.Save()
        $result | ConvertTo-Json -Compress -Depth 6
    }
    finally {
        $wb.Close($true)
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($wb) | Out-Null
    }
}
finally {
    $excel.Quit()
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
}
"""

    with tempfile.TemporaryDirectory() as tmp_dir:
        script_path = Path(tmp_dir) / "excel_com_master_index_helper.ps1"
        payload_path = Path(tmp_dir) / "payload.json"
        script_path.write_text(script, encoding="utf-8")
        payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        env = os.environ.copy()
        env["MASTER_INDEX_HELPER_PAYLOAD"] = str(payload_path)
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            env=env,
            check=True,
        )
    stdout = result.stdout.strip()
    if not stdout:
        return {"status": "ok"}
    return json.loads(stdout)


def _ensure_database_asset_sheet_excel_com(workbook_path: Path) -> None:
    _run_excel_com_sheet_operation(workbook_path=workbook_path, operation="ensure")


def _rows_to_dicts(ws) -> list[dict[str, Any]]:
    headers = _sheet_headers(ws)
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if all(value in (None, "") for value in row):
            continue
        record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        if "latest_review_date" in record:
            record["latest_review_date"] = _normalize_date_like_text(record["latest_review_date"])
        records.append(record)
    return records


def read_database_asset_records(
    workbook_path: str | Path,
    std_variable_id: str | None = None,
    database_id: str | None = None,
) -> list[dict[str, Any]]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Master index workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if DATABASE_ASSET_SHEET not in wb.sheetnames:
            return []
        ws = wb[DATABASE_ASSET_SHEET]
        records = _rows_to_dicts(ws)
    finally:
        wb.close()

    filtered: list[dict[str, Any]] = []
    for record in records:
        if std_variable_id and record.get("std_variable_id") != std_variable_id:
            continue
        if database_id and record.get("database_id") != database_id:
            continue
        filtered.append(record)
    return filtered


def read_variable_catalog_records(
    workbook_path: str | Path,
    std_variable_id: str | None = None,
) -> list[dict[str, Any]]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Master index workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if VARIABLE_CATALOG_SHEET not in wb.sheetnames:
            return []
        ws = wb[VARIABLE_CATALOG_SHEET]
        records = _rows_to_dicts(ws)
    finally:
        wb.close()

    if std_variable_id is None:
        return records
    return [record for record in records if record.get("std_variable_id") == std_variable_id]


def upsert_database_asset_record(
    workbook_path: str | Path,
    record: dict[str, Any],
) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        _ensure_database_asset_sheet_openpyxl(workbook_path)

    if _prefer_excel_com_backend():
        try:
            return _run_excel_com_sheet_operation(
                workbook_path=workbook_path,
                operation="upsert",
                record=record,
            )
        except Exception:
            if os.getenv("CLINICAL_DATABASE_MASTER_INDEX_BACKEND", "auto").strip().lower() == "excel_com":
                raise

    return _upsert_database_asset_record_openpyxl(workbook_path, record)


def _upsert_database_asset_record_openpyxl(
    workbook_path: Path,
    record: dict[str, Any],
) -> dict[str, Any]:
    missing = sorted(REQUIRED_DATABASE_ASSET_KEYS - set(record))
    if missing:
        raise ValueError(f"Missing required database-asset keys: {missing}")

    ensure_database_asset_sheet(workbook_path)
    wb = load_workbook(workbook_path)
    try:
        ws = wb[DATABASE_ASSET_SHEET]
        headers = _sheet_headers(ws)
        header_to_col = {header: idx for idx, header in enumerate(headers, start=1)}

        target_row = None
        for row_index in range(2, ws.max_row + 1):
            std_id = ws.cell(row=row_index, column=header_to_col["std_variable_id"]).value
            db_id = ws.cell(row=row_index, column=header_to_col["database_id"]).value
            if std_id == record["std_variable_id"] and db_id == record["database_id"]:
                target_row = row_index
                break

        action = "updated" if target_row is not None else "inserted"
        if target_row is None:
            target_row = ws.max_row + 1

        for header in DATABASE_ASSET_HEADERS:
            value = record.get(header, "")
            ws.cell(row=target_row, column=header_to_col[header], value=_normalize_scalar(value))

        wb.save(workbook_path)
        return {
            "action": action,
            "row_index": target_row,
            "std_variable_id": record["std_variable_id"],
            "database_id": record["database_id"],
            "workbook_path": str(workbook_path),
        }
    finally:
        wb.close()


def _load_record_payload(args) -> dict[str, Any]:
    if args.record_json:
        return json.loads(Path(args.record_json).read_text(encoding="utf-8-sig"))
    if args.record_inline:
        return json.loads(args.record_inline)
    raise ValueError("Either --record-json or --record-inline is required for upsert.")


def _print_json(payload: Any) -> None:
    print(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Read and write the Layer 5 global master index database-asset registry."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    ensure_parser = subparsers.add_parser("ensure-sheet", help="Ensure the database-asset sheet exists.")
    ensure_parser.add_argument("--workbook", required=True, help="Path to Layer5_StdVariable_MasterIndex.xlsx")

    read_parser = subparsers.add_parser("read", help="Read rows from std_variable_database_assets.")
    read_parser.add_argument("--workbook", required=True, help="Path to Layer5_StdVariable_MasterIndex.xlsx")
    read_parser.add_argument("--std-variable-id", dest="std_variable_id")
    read_parser.add_argument("--database-id", dest="database_id")

    upsert_parser = subparsers.add_parser("upsert", help="Insert or update one database-asset record.")
    upsert_parser.add_argument("--workbook", required=True, help="Path to Layer5_StdVariable_MasterIndex.xlsx")
    upsert_parser.add_argument("--record-json", help="Path to a JSON file containing one record object.")
    upsert_parser.add_argument("--record-inline", help="Inline JSON object for one record.")

    args = parser.parse_args()

    if args.command == "ensure-sheet":
        ensure_database_asset_sheet(args.workbook)
        _print_json({"status": "ok", "sheet": DATABASE_ASSET_SHEET, "workbook": str(Path(args.workbook))})
        return

    if args.command == "read":
        records = read_database_asset_records(
            workbook_path=args.workbook,
            std_variable_id=args.std_variable_id,
            database_id=args.database_id,
        )
        _print_json(records)
        return

    if args.command == "upsert":
        payload = _load_record_payload(args)
        result = upsert_database_asset_record(args.workbook, payload)
        _print_json(result)
        return


if __name__ == "__main__":
    main()
